import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
from typing import Dict, Optional, Tuple

class ExcelReader:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
    
    def open_workbook(self) -> None:
        """Opens the Excel workbook"""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
        except Exception as e:
            raise Exception(f"Error opening workbook: {str(e)}")
    
    def get_sheet(self, sheet_name: str):
        """Returns specified worksheet"""
        if not self.workbook:
            self.open_workbook()
        return self.workbook[sheet_name]
    
    def create_sheet(self, sheet_name: str):
        """Creates a new sheet in the workbook"""
        if not self.workbook:
            self.open_workbook()
        if sheet_name in self.workbook.sheetnames:
            self.workbook.remove(self.workbook[sheet_name])
        return self.workbook.create_sheet(sheet_name)
    
    def save_workbook(self):
        """Saves the workbook"""
        if self.workbook:
            self.workbook.save(self.file_path)

class ProductPortfolio:
    def __init__(self, excel_reader: ExcelReader):
        self.excel_reader = excel_reader
        self.products: Dict = {}
        self.components: Dict = {}
    
    def process_portfolio(self) -> None:
        """Process CarteiraSAP and calculate required components"""
        carteira_sheet = self.excel_reader.get_sheet("CarteiraSAP")
        bom_sheet = self.excel_reader.get_sheet("BOM SAP")
        
        # Read products from CarteiraSAP (starting from row 2)
        for row in carteira_sheet.iter_rows(min_row=2, values_only=True):
            product_code = row[0]  # Column A: Product Code
            description = row[1]   # Column B: Description
            quantity = row[2]      # Column C: Total Quantity
            due_date = row[3]      # Column D: Due Date
            
            self.products[product_code] = {
                "description": description,
                "quantity": quantity,
                "due_date": due_date
            }
            
            # Find all components in BOM SAP
            self._process_bom_components(bom_sheet, product_code, quantity)
    
    def _process_bom_components(self, bom_sheet, product_code: str, quantity: float) -> None:
        """Process BOM components for a given product"""
        for row in bom_sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == product_code:  # Match product code in BOM
                component_code = row[3]    # Component code
                component_desc = row[4]    # Component description
                factor = float(row[5] or 0)  # Quantity factor
                
                total_qty = quantity * factor
                
                if component_code in self.components:
                    self.components[component_code]["quantity"] += total_qty
                else:
                    self.components[component_code] = {
                        "description": component_desc,
                        "quantity": total_qty
                    }

def main():
    file_path = r"C:\scripts\job_code_files\automações\auto_atraso\results.xlsx"
    
    # Initialize classes
    excel_reader = ExcelReader(file_path)
    portfolio = ProductPortfolio(excel_reader)
    
    # Process product portfolio
    portfolio.process_portfolio()
    
    # Get products and their components from BOM
    bom_sheet = excel_reader.get_sheet("BOM SAP")
    product_components = {}
      # Group components by product
    for row in bom_sheet.iter_rows(min_row=2, values_only=True):
        product_code = row[0]
        if product_code in portfolio.products:  # Only process products in our portfolio
            component_code = row[3]
            component_desc = row[4]
            factor = float(row[5] or 0)
            
            if product_code not in product_components:
                product_components[product_code] = []
            
            # Calculate quantity needed for this specific product
            product_quantity = portfolio.products[product_code]["quantity"]
            component_quantity = product_quantity * factor
            
            product_components[product_code].append({
                "code": component_code,
                "description": component_desc,
                "factor": factor,
                "total_quantity": component_quantity
            })
    
    # Create new sheet for results
    results_sheet = excel_reader.create_sheet("Processed BOM")
    current_row = 1
    
    # Write headers
    headers = ["Produto", "Descrição", "Quantidade Total", "Data de Atraso", "Factor"]
    for col, header in enumerate(headers, 1):
        results_sheet.cell(row=current_row, column=col, value=header)
        # Set column width
        results_sheet.column_dimensions[get_column_letter(col)].width = 20
    
    current_row += 1
    
    # Write results in flat structure
    for product_code, product_data in portfolio.products.items():
        # Write product information
        results_sheet.cell(row=current_row, column=1, value=product_code)
        results_sheet.cell(row=current_row, column=2, value=product_data['description'])
        results_sheet.cell(row=current_row, column=3, value=product_data['quantity'])
        results_sheet.cell(row=current_row, column=4, value=product_data['due_date'])
        
        current_row += 1
          # Write components
        if product_code in product_components:
            for comp in product_components[product_code]:
                results_sheet.cell(row=current_row, column=1, value=comp['code'])
                results_sheet.cell(row=current_row, column=2, value=comp['description'])
                results_sheet.cell(row=current_row, column=3, value=comp['total_quantity'])
                results_sheet.cell(row=current_row, column=5, value=comp['factor'])
                current_row += 1
                
        # Add empty row between main products
        current_row += 1
    
    # Save the workbook
    excel_reader.save_workbook()
    print(f"Results have been written to sheet 'Processed BOM' in {file_path}")

if __name__ == "__main__":
    main()