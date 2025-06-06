import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

def corrigir_descricao_linhas(linhas):
    """
    Corrige as descrições inválidas nas linhas da planilha, removendo todos os caracteres antes de 'CB' ou 'FIO'.
    """
    linhas_corrigidas = []
    for linha in linhas:
        produto = str(linha[4] or "").strip()  # Coluna C: Produto (índice 2)
        descricao = str(linha[5] or "").strip()  # Coluna D: Descrição (índice 3)

        # Procurar "CB" ou "FIO" na descrição
        if "CB" in descricao:
            descricao = descricao[descricao.index("CB"):]
        elif "FIO" in descricao:
            descricao = descricao[descricao.index("FIO"):]
        else:
            print(f"Não foi possível corrigir a descrição para o produto: {produto}")

        # Substituir a linha corrigida
        linha_corrigida = list(linha)
        linha_corrigida[5] = descricao  # Atualizar a descrição corrigida
        linhas_corrigidas.append(tuple(linha_corrigida))
    return linhas_corrigidas

def main():
    # Abrir o arquivo Excel
    caminho_arquivo = r"C:\scripts\job_code_files\automações\auto_atraso\results.xlsx"
    try:
        workbook = openpyxl.load_workbook(caminho_arquivo)
    except FileNotFoundError:
        print("Arquivo não encontrado.")
        return

    # Definir as planilhas
    try:
        ws_origem = workbook["Planilha1"]
    except KeyError:
        print("Planilha 'AtrasoSap' não encontrada.")
        return

    if "CarteiraSAP" in workbook.sheetnames:
        workbook.remove(workbook["CarteiraSAP"])
    ws_destino = workbook.create_sheet("CarteiraSAP")

    # Ler as linhas da planilha de origem
    linhas = list(ws_origem.iter_rows(min_row=2, max_row=ws_origem.max_row, values_only=True))

    # Corrigir as descrições nas linhas
    linhas_corrigidas = corrigir_descricao_linhas(linhas)

    # Criar um dicionário para agrupar os valores
    agrupamento = {}

    # Percorrer as linhas corrigidas para realizar o agrupamento
    for linha in linhas_corrigidas:
        produto = str(linha[4] or "").strip()  # Coluna C: Produto (índice 2)
        descricao = str(linha[5] or "").strip()  # Coluna D: Descrição (índice 3)
        saldo = linha[9] if isinstance(linha[9], (int, float)) else 0  # Coluna J: Quantidade Pedida (índice 9)
        data_entrega = linha[2] if isinstance(linha[2], datetime) else datetime.today()  # Coluna E: Data de Entrega (índice 4)

        chave = f"{produto}|{descricao}"

        if chave in agrupamento:
            agrupamento[chave]["saldo"] += saldo
            if data_entrega < agrupamento[chave]["data_entrega"]:
                agrupamento[chave]["data_entrega"] = data_entrega
        else:
            agrupamento[chave] = {
                "produto": produto,
                "descricao": descricao,
                "saldo": saldo,
                "data_entrega": data_entrega
            }

    # Preencher a planilha de destino
    ws_destino.append(["Produto", "Descrição", "Quantidade Total", "Data de Atraso"])
    for dados in agrupamento.values():
        ws_destino.append([
            dados["produto"],
            dados["descricao"],
            dados["saldo"],
            dados["data_entrega"].strftime("%d-%m-%Y")
        ])

    # Ajustar a largura das colunas
    for col in ws_destino.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws_destino.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # Salvar o arquivo
    workbook.save(caminho_arquivo)
    print("Agrupamento e soma concluídos na planilha 'CarteiraSAP'.")

main()